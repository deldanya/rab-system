"""
Parser Excel fleksibel untuk membaca file RAB dengan format berbeda-beda.
Strategi: deteksi otomatis posisi kolom berdasarkan keyword, lalu parsing baris.
"""
import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional, Tuple, Any
from pathlib import Path


# ─── Keyword untuk deteksi kolom ──────────────────────────────────────────────

KEYWORD_NOMOR = ["no", "no.", "nomor", "urut", "nomer"]
KEYWORD_URAIAN = ["uraian", "pekerjaan", "item", "keterangan", "deskripsi", "nama pekerjaan", "uraian pekerjaan"]
KEYWORD_SATUAN = ["satuan", "sat", "unit", "units"]
KEYWORD_VOLUME = ["volume", "vol", "qty", "kuantitas", "jumlah pekerjaan", "quantity"]
KEYWORD_HARGA_SATUAN = ["harga satuan", "harga", "unit price", "satuan harga", "h.satuan", "h satuan", "price"]
KEYWORD_JUMLAH = ["jumlah", "total", "amount", "nilai", "sub total", "jumlah harga"]

# Keyword header info proyek
KEYWORD_PROYEK = ["nama proyek", "proyek", "pekerjaan", "nama kegiatan", "kegiatan"]
KEYWORD_LOKASI = ["lokasi", "tempat", "alamat", "lokasi kegiatan", "kota", "kabupaten"]
KEYWORD_TAHUN = ["tahun", "tahun anggaran", "t.a.", "t.a"]


def bersihkan_angka(nilai: Any) -> Optional[float]:
    """
    Konversi nilai apapun ke float, termasuk format "Rp 1.000.000,00" atau "1,000,000.00".
    Kembalikan None jika tidak bisa dikonversi.
    """
    if nilai is None:
        return None
    if isinstance(nilai, (int, float)):
        if pd.isna(nilai):
            return None
        return float(nilai)

    teks = str(nilai).strip()
    if not teks or teks.lower() in ["-", "n/a", "na", ""]:
        return None

    # Hapus simbol mata uang dan spasi
    teks = re.sub(r"[Rp\s]", "", teks, flags=re.IGNORECASE)
    teks = teks.replace("IDR", "").replace("idr", "")

    # Deteksi format Indonesia (titik=ribuan, koma=desimal) vs format internasional
    jumlah_titik = teks.count(".")
    jumlah_koma = teks.count(",")

    if jumlah_koma > 0 and jumlah_titik > 0:
        # Misal: "1.000.000,50" → format Indonesia
        if teks.rfind(",") > teks.rfind("."):
            teks = teks.replace(".", "").replace(",", ".")
        else:
            # "1,000,000.50" → format internasional
            teks = teks.replace(",", "")
    elif jumlah_koma > 0 and jumlah_titik == 0:
        # Mungkin "1,000" (ribuan) atau "1,5" (desimal)
        parts = teks.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            # Kemungkinan desimal: "1,50" → "1.50"
            teks = teks.replace(",", ".")
        else:
            # Kemungkinan ribuan Indonesia: "1,000,000" → "1000000"
            teks = teks.replace(",", "")
    elif jumlah_titik > 1:
        # "1.000.000" → format ribuan Indonesia
        teks = teks.replace(".", "")

    teks = re.sub(r"[^\d\.\-]", "", teks)

    try:
        return float(teks) if teks else None
    except (ValueError, TypeError):
        return None


def cocokkan_keyword(teks_header: str, keywords: List[str]) -> bool:
    """Cek apakah teks header mengandung salah satu keyword."""
    if not teks_header:
        return False
    teks = str(teks_header).lower().strip()
    return any(kw in teks for kw in keywords)


def deteksi_baris_header(sheet, max_scan: int = 30) -> Tuple[Optional[int], Dict[str, int]]:
    """
    Scan baris-baris awal untuk menemukan baris header tabel BOQ.
    Kembalikan (nomor_baris_header, mapping_kolom).
    """
    for row_idx in range(1, min(max_scan, sheet.max_row) + 1):
        baris_teks = []
        for col_idx in range(1, min(20, sheet.max_column) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            nilai = cell.value
            if nilai is None and cell.coordinate in sheet.merged_cells:
                # Cari nilai dari sel merged
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        nilai = sheet.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        ).value
                        break
            baris_teks.append(str(nilai).lower().strip() if nilai else "")

        teks_gabung = " ".join(baris_teks)

        # Baris header harus mengandung minimal 3 keyword
        skor = 0
        if any(k in teks_gabung for k in KEYWORD_URAIAN):
            skor += 2
        if any(k in teks_gabung for k in KEYWORD_SATUAN):
            skor += 2
        if any(k in teks_gabung for k in KEYWORD_VOLUME):
            skor += 2
        if any(k in teks_gabung for k in KEYWORD_HARGA_SATUAN):
            skor += 2
        if any(k in teks_gabung for k in KEYWORD_JUMLAH):
            skor += 1
        if any(k in teks_gabung for k in KEYWORD_NOMOR):
            skor += 1

        if skor >= 5:
            # Ditemukan! Map kolom
            mapping = {}
            for col_idx, teks in enumerate(baris_teks, start=1):
                if not teks:
                    continue
                if cocokkan_keyword(teks, KEYWORD_NOMOR) and "nomor" not in mapping:
                    mapping["nomor"] = col_idx
                elif cocokkan_keyword(teks, KEYWORD_URAIAN) and "uraian" not in mapping:
                    mapping["uraian"] = col_idx
                elif cocokkan_keyword(teks, KEYWORD_SATUAN) and "satuan" not in mapping:
                    mapping["satuan"] = col_idx
                elif cocokkan_keyword(teks, KEYWORD_VOLUME) and "volume" not in mapping:
                    mapping["volume"] = col_idx
                elif cocokkan_keyword(teks, KEYWORD_HARGA_SATUAN) and "harga_satuan" not in mapping:
                    mapping["harga_satuan"] = col_idx
                elif cocokkan_keyword(teks, KEYWORD_JUMLAH) and "jumlah" not in mapping:
                    mapping["jumlah"] = col_idx

            if "uraian" in mapping:
                return row_idx, mapping

    return None, {}


def ekstrak_metadata_proyek(sheet, baris_header: int) -> Dict[str, Any]:
    """
    Cari metadata proyek (nama, lokasi, tahun) di baris sebelum header tabel.
    """
    metadata = {"nama_proyek": None, "lokasi": None, "tahun": None}

    for row_idx in range(1, baris_header):
        for col_idx in range(1, min(10, sheet.max_column) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            nilai = str(cell.value).strip() if cell.value else ""
            if not nilai or nilai.lower() == "none":
                continue

            nilai_lower = nilai.lower()

            if any(k in nilai_lower for k in KEYWORD_PROYEK) and not metadata["nama_proyek"]:
                # Ambil nilai di kolom sebelah kanan
                for offset in range(1, 5):
                    next_val = sheet.cell(row=row_idx, column=col_idx + offset).value
                    if next_val and str(next_val).strip():
                        metadata["nama_proyek"] = str(next_val).strip()
                        break
                # Atau ambil teks setelah tanda ":"
                if not metadata["nama_proyek"] and ":" in nilai:
                    bagian = nilai.split(":", 1)
                    if len(bagian) > 1 and bagian[1].strip():
                        metadata["nama_proyek"] = bagian[1].strip()

            if any(k in nilai_lower for k in KEYWORD_LOKASI) and not metadata["lokasi"]:
                for offset in range(1, 5):
                    next_val = sheet.cell(row=row_idx, column=col_idx + offset).value
                    if next_val and str(next_val).strip():
                        metadata["lokasi"] = str(next_val).strip()
                        break
                if not metadata["lokasi"] and ":" in nilai:
                    bagian = nilai.split(":", 1)
                    if len(bagian) > 1 and bagian[1].strip():
                        metadata["lokasi"] = bagian[1].strip()

            if any(k in nilai_lower for k in KEYWORD_TAHUN) and not metadata["tahun"]:
                for offset in range(1, 5):
                    next_val = sheet.cell(row=row_idx, column=col_idx + offset).value
                    angka = bersihkan_angka(next_val)
                    if angka and 2000 <= angka <= 2100:
                        metadata["tahun"] = int(angka)
                        break
                if not metadata["tahun"] and ":" in nilai:
                    bagian = nilai.split(":", 1)
                    angka = bersihkan_angka(bagian[1]) if len(bagian) > 1 else None
                    if angka and 2000 <= angka <= 2100:
                        metadata["tahun"] = int(angka)

            # Deteksi tahun langsung dari nilai sel (4 digit angka)
            if not metadata["tahun"]:
                match = re.search(r'\b(20\d{2})\b', nilai)
                if match:
                    metadata["tahun"] = int(match.group(1))

    return metadata


def apakah_baris_divisi(nilai_uraian: str) -> bool:
    """
    Deteksi apakah baris adalah header divisi/kelompok pekerjaan
    (bukan item pekerjaan konkret).
    """
    if not nilai_uraian:
        return False
    teks = str(nilai_uraian).strip()

    # Pola umum divisi: huruf Romawi atau huruf besar diikuti titik
    if re.match(r'^[IVX]+[\.\s]', teks) or re.match(r'^[A-Z][\.\s]', teks):
        return True

    # Semua huruf besar tanpa angka = kemungkinan header
    if teks.isupper() and len(teks) > 3 and not re.search(r'\d', teks):
        return True

    # Keyword umum divisi
    keyword_divisi = [
        "pekerjaan persiapan", "pekerjaan tanah", "pekerjaan pondasi",
        "pekerjaan beton", "pekerjaan baja", "pekerjaan kayu",
        "pekerjaan atap", "pekerjaan plesteran", "pekerjaan finishing",
        "pekerjaan mekanikal", "pekerjaan elektrikal", "pekerjaan sanitasi",
        "pekerjaan pasangan", "pekerjaan pengecatan", "pekerjaan penutup lantai",
    ]
    teks_lower = teks.lower()
    if any(kd in teks_lower for kd in keyword_divisi) and len(teks.split()) <= 6:
        return True

    return False


def parse_excel_rab(path_file: str) -> Dict:
    """
    Fungsi utama: parse file Excel RAB.
    Kembalikan dict berisi metadata proyek dan daftar item BOQ.
    """
    hasil = {
        "sukses": False,
        "pesan": "",
        "metadata": {"nama_proyek": None, "lokasi": None, "tahun": None},
        "items": [],
        "total_baris": 0,
        "total_item_valid": 0,
        "total_divisi": 0,
    }

    try:
        wb = openpyxl.load_workbook(path_file, data_only=True)
    except Exception as e:
        hasil["pesan"] = f"Gagal membuka file Excel: {str(e)}"
        return hasil

    # Coba setiap sheet, ambil yang pertama valid
    sheet_terpilih = None
    baris_header = None
    mapping_kolom = {}

    for nama_sheet in wb.sheetnames:
        sheet = wb[nama_sheet]
        if sheet.max_row < 3:
            continue
        br, mk = deteksi_baris_header(sheet)
        if br and "uraian" in mk:
            sheet_terpilih = sheet
            baris_header = br
            mapping_kolom = mk
            break

    if not sheet_terpilih:
        hasil["pesan"] = (
            "Tidak dapat mendeteksi format tabel RAB. "
            "Pastikan file memiliki kolom: Uraian Pekerjaan, Satuan, Volume, Harga Satuan, Jumlah. "
            "Gunakan template standar yang tersedia."
        )
        return hasil

    # Ekstrak metadata dari baris sebelum tabel
    metadata = ekstrak_metadata_proyek(sheet_terpilih, baris_header)
    hasil["metadata"] = metadata

    # Parse baris data
    items = []
    divisi_sekarang = None
    total_baris = 0
    total_item_valid = 0
    total_divisi = 0

    for row_idx in range(baris_header + 1, sheet_terpilih.max_row + 1):
        # Ambil nilai dari tiap kolom yang sudah dipetakan
        def ambil_nilai(nama_kol):
            if nama_kol not in mapping_kolom:
                return None
            cell = sheet_terpilih.cell(row=row_idx, column=mapping_kolom[nama_kol])
            val = cell.value
            # Tangani merged cells
            if val is None:
                for mr in sheet_terpilih.merged_cells.ranges:
                    if cell.coordinate in mr:
                        val = sheet_terpilih.cell(row=mr.min_row, column=mr.min_col).value
                        break
            return val

        uraian_raw = ambil_nilai("uraian")
        nomor_raw = ambil_nilai("nomor")
        satuan_raw = ambil_nilai("satuan")
        volume_raw = ambil_nilai("volume")
        harga_sat_raw = ambil_nilai("harga_satuan")
        jumlah_raw = ambil_nilai("jumlah")

        if not uraian_raw or str(uraian_raw).strip() == "" or str(uraian_raw).strip().lower() == "none":
            continue

        uraian = str(uraian_raw).strip()
        total_baris += 1

        # Cek apakah baris divisi
        nomor_str = str(nomor_raw).strip() if nomor_raw else ""
        if apakah_baris_divisi(uraian) or (nomor_str and not re.search(r'\d', nomor_str) and len(uraian) > 5):
            divisi_sekarang = uraian
            items.append({
                "nomor_urut": nomor_str or None,
                "divisi": None,
                "is_divisi": True,
                "uraian_pekerjaan": uraian,
                "satuan": None,
                "volume": None,
                "harga_satuan": None,
                "jumlah": None,
                "peringatan": None,
            })
            total_divisi += 1
            continue

        # Baris item biasa
        volume = bersihkan_angka(volume_raw)
        harga_satuan = bersihkan_angka(harga_sat_raw)
        jumlah = bersihkan_angka(jumlah_raw)

        # Hitung jumlah jika tidak ada tapi ada volume & harga
        if jumlah is None and volume is not None and harga_satuan is not None:
            jumlah = volume * harga_satuan

        # Validasi silang: peringatkan jika jumlah tidak cocok
        peringatan = None
        if (jumlah is not None and volume is not None and harga_satuan is not None):
            selisih = abs(jumlah - (volume * harga_satuan))
            if selisih > 1 and selisih / (jumlah + 0.01) > 0.01:
                peringatan = (
                    f"Jumlah di Excel ({jumlah:,.0f}) tidak cocok dengan "
                    f"Volume × Harga Satuan ({volume * harga_satuan:,.0f}). "
                    "Nilai jumlah akan dihitung ulang."
                )
                jumlah = volume * harga_satuan

        # Lewati baris yang tidak punya nilai apapun
        if volume is None and harga_satuan is None and jumlah is None:
            peringatan_kosong = "Baris tanpa nilai volume/harga - dilewati"
            # Tetap simpan jika ada uraian yang berarti (jangan abaikan)
            if not satuan_raw:
                continue

        satuan = str(satuan_raw).strip() if satuan_raw else None

        items.append({
            "nomor_urut": nomor_str or None,
            "divisi": divisi_sekarang,
            "is_divisi": False,
            "uraian_pekerjaan": uraian,
            "satuan": satuan,
            "volume": volume,
            "harga_satuan": harga_satuan,
            "jumlah": jumlah,
            "peringatan": peringatan,
        })
        total_item_valid += 1

    hasil["sukses"] = True
    hasil["items"] = items
    hasil["total_baris"] = total_baris
    hasil["total_item_valid"] = total_item_valid
    hasil["total_divisi"] = total_divisi
    hasil["pesan"] = (
        f"Berhasil membaca {total_item_valid} item pekerjaan "
        f"dan {total_divisi} kelompok pekerjaan."
    )
    return hasil
