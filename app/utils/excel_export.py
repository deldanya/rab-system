"""
Ekspor RAB ke format Excel profesional siap cetak.
Menggunakan openpyxl untuk kontrol penuh atas formatting.
"""
import os
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as num_fmt
)
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent.parent
EXPORT_DIR = BASE_DIR / "data" / "exports"
TEMPLATE_DIR = BASE_DIR / "data" / "templates"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

# ─── Style konstanta ──────────────────────────────────────────────────────────

WARNA_BIRU_TUA = "1F4E79"
WARNA_BIRU_MUDA = "BDD7EE"
WARNA_KUNING = "FFF2CC"
WARNA_HIJAU = "E2EFDA"
WARNA_MERAH_MUDA = "FCE4D6"
WARNA_ABU = "F2F2F2"
WARNA_PUTIH = "FFFFFF"

BORDER_TIPIS = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BORDER_MEDIUM = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)

FORMAT_RUPIAH = '#,##0'
FORMAT_RUPIAH_DES = '#,##0.00'
FORMAT_PERSEN = '0.00%'


def terapkan_border(ws, baris_min, kolom_min, baris_max, kolom_max, border=BORDER_TIPIS):
    for row in ws.iter_rows(min_row=baris_min, max_row=baris_max,
                             min_col=kolom_min, max_col=kolom_max):
        for cell in row:
            cell.border = border


def set_cell(ws, baris, kolom, nilai, bold=False, center=False, fill_hex=None,
             font_size=10, format_angka=None, wrap=False, border=None, italic=False):
    cell = ws.cell(row=baris, column=kolom, value=nilai)
    cell.font = Font(name="Calibri", size=font_size, bold=bold, italic=italic)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap,
    )
    if fill_hex:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    if format_angka:
        cell.number_format = format_angka
    if border:
        cell.border = border
    return cell


def buat_excel_rab(data_rab: Dict[str, Any]) -> str:
    """
    Buat file Excel RAB profesional.
    data_rab: {nama_proyek, lokasi, tahun, deskripsi, ppn_persen, items, subtotal, ppn_nominal, total}
    Kembalikan path file yang dibuat.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAB"

    # ── Lebar kolom ──
    ws.column_dimensions["A"].width = 6   # No
    ws.column_dimensions["B"].width = 45  # Uraian Pekerjaan
    ws.column_dimensions["C"].width = 10  # Satuan
    ws.column_dimensions["D"].width = 12  # Volume
    ws.column_dimensions["E"].width = 18  # Harga Satuan
    ws.column_dimensions["F"].width = 20  # Jumlah

    baris = 1

    # ── Judul ──
    ws.merge_cells(f"A{baris}:F{baris}")
    set_cell(ws, baris, 1, "RENCANA ANGGARAN BIAYA (RAB)",
             bold=True, center=True, fill_hex=WARNA_BIRU_TUA,
             font_size=14)
    ws.cell(row=baris, column=1).font = Font(
        name="Calibri", size=14, bold=True, color="FFFFFF"
    )
    ws.row_dimensions[baris].height = 28
    baris += 1

    # ── Info Proyek ──
    info_rows = [
        ("Nama Proyek", data_rab.get("nama_proyek", "-")),
        ("Lokasi", data_rab.get("lokasi", "-")),
        ("Tahun Anggaran", str(data_rab.get("tahun", "-"))),
        ("Tanggal Dibuat", datetime.now().strftime("%d %B %Y")),
    ]
    if data_rab.get("deskripsi"):
        info_rows.append(("Keterangan", data_rab["deskripsi"]))

    for label, nilai in info_rows:
        ws.merge_cells(f"A{baris}:B{baris}")
        set_cell(ws, baris, 1, label, bold=True, fill_hex=WARNA_ABU, border=BORDER_TIPIS)
        ws.merge_cells(f"C{baris}:F{baris}")
        set_cell(ws, baris, 3, nilai, fill_hex=WARNA_PUTIH, border=BORDER_TIPIS)
        baris += 1

    baris += 1  # Baris kosong

    # ── Header Tabel ──
    headers = ["No", "Uraian Pekerjaan", "Sat", "Volume", "Harga Satuan (Rp)", "Jumlah (Rp)"]
    for col_idx, header in enumerate(headers, start=1):
        set_cell(ws, baris, col_idx, header,
                 bold=True, center=True,
                 fill_hex=WARNA_BIRU_TUA,
                 font_size=10, border=BORDER_TIPIS)
        ws.cell(row=baris, column=col_idx).font = Font(
            name="Calibri", size=10, bold=True, color="FFFFFF"
        )
    ws.row_dimensions[baris].height = 20
    baris_header_tabel = baris
    baris += 1

    # ── Isi Item BOQ ──
    items = data_rab.get("items", [])
    nomor_item = 0
    subtotal_per_divisi = {}
    divisi_sekarang = None

    baris_awal_data = baris

    for item in items:
        is_divisi = item.get("is_divisi", False)

        if is_divisi:
            # Baris divisi / header kelompok
            divisi_sekarang = item.get("uraian_pekerjaan", "")
            ws.merge_cells(f"A{baris}:F{baris}")
            cell = ws.cell(row=baris, column=1)
            cell.value = divisi_sekarang
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor=WARNA_BIRU_MUDA)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER_TIPIS
            ws.row_dimensions[baris].height = 18
        else:
            nomor_item += 1
            jumlah = item.get("jumlah") or (
                (item.get("volume") or 0) * (item.get("harga_satuan") or 0)
            )

            # Kolom A: nomor
            set_cell(ws, baris, 1, item.get("nomor_urut") or nomor_item,
                     center=True, border=BORDER_TIPIS)
            # Kolom B: uraian
            set_cell(ws, baris, 2, item.get("uraian_pekerjaan", ""),
                     wrap=True, border=BORDER_TIPIS)
            # Kolom C: satuan
            set_cell(ws, baris, 3, item.get("satuan", ""),
                     center=True, border=BORDER_TIPIS)
            # Kolom D: volume
            set_cell(ws, baris, 4, item.get("volume"),
                     center=True, format_angka=FORMAT_RUPIAH_DES, border=BORDER_TIPIS)
            # Kolom E: harga satuan
            set_cell(ws, baris, 5, item.get("harga_satuan"),
                     format_angka=FORMAT_RUPIAH, border=BORDER_TIPIS)
            ws.cell(row=baris, column=5).alignment = Alignment(horizontal="right", vertical="center")
            # Kolom F: jumlah
            set_cell(ws, baris, 6, jumlah,
                     fill_hex=WARNA_HIJAU,
                     format_angka=FORMAT_RUPIAH, border=BORDER_TIPIS)
            ws.cell(row=baris, column=6).alignment = Alignment(horizontal="right", vertical="center")

        baris += 1

    # ── Rekapitulasi ──
    baris += 1
    ws.merge_cells(f"A{baris}:E{baris}")
    set_cell(ws, baris, 1, "SUBTOTAL PEKERJAAN",
             bold=True, center=True, fill_hex=WARNA_KUNING, border=BORDER_TIPIS)
    subtotal = data_rab.get("subtotal", 0)
    set_cell(ws, baris, 6, subtotal,
             bold=True, fill_hex=WARNA_KUNING,
             format_angka=FORMAT_RUPIAH, border=BORDER_TIPIS)
    ws.cell(row=baris, column=6).alignment = Alignment(horizontal="right", vertical="center")
    baris += 1

    ppn_persen = data_rab.get("ppn_persen", 11.0)
    ppn_nominal = data_rab.get("ppn_nominal", 0)
    ws.merge_cells(f"A{baris}:E{baris}")
    set_cell(ws, baris, 1, f"PPN {ppn_persen:.0f}%",
             bold=True, center=True, fill_hex=WARNA_MERAH_MUDA, border=BORDER_TIPIS)
    set_cell(ws, baris, 6, ppn_nominal,
             bold=True, fill_hex=WARNA_MERAH_MUDA,
             format_angka=FORMAT_RUPIAH, border=BORDER_TIPIS)
    ws.cell(row=baris, column=6).alignment = Alignment(horizontal="right", vertical="center")
    baris += 1

    total = data_rab.get("total", 0)
    ws.merge_cells(f"A{baris}:E{baris}")
    set_cell(ws, baris, 1, "TOTAL BIAYA",
             bold=True, center=True, fill_hex=WARNA_BIRU_TUA,
             font_size=11, border=BORDER_TIPIS)
    ws.cell(row=baris, column=1).font = Font(
        name="Calibri", size=11, bold=True, color="FFFFFF"
    )
    set_cell(ws, baris, 6, total,
             bold=True, fill_hex=WARNA_BIRU_TUA,
             font_size=11, format_angka=FORMAT_RUPIAH, border=BORDER_TIPIS)
    ws.cell(row=baris, column=6).font = Font(
        name="Calibri", size=11, bold=True, color="FFFFFF"
    )
    ws.cell(row=baris, column=6).alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[baris].height = 22
    baris += 1

    # ── Catatan kaki ──
    baris += 1
    ws.merge_cells(f"A{baris}:F{baris}")
    set_cell(ws, baris, 1,
             f"Dokumen ini dibuat otomatis oleh Sistem RAB pada {datetime.now().strftime('%d/%m/%Y %H:%M')}",
             italic=True, font_size=8)

    # ── Freeze panes di baris header ──
    ws.freeze_panes = ws.cell(row=baris_header_tabel + 1, column=1)

    # ── Simpan file ──
    nama_aman = "".join(c for c in data_rab.get("nama_proyek", "RAB") if c.isalnum() or c in " -_")
    nama_file = f"RAB_{nama_aman}_{data_rab.get('tahun', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path_output = EXPORT_DIR / nama_file

    wb.save(str(path_output))
    return str(path_output)


def buat_template_excel() -> str:
    """
    Buat file template Excel yang dijamin bisa diparsing sistem.
    Format: metadata 3 baris atas, baris 4 kosong, header baris 5, data mulai baris 6.
    Parser mendeteksi header dengan keyword — pastikan nama kolom tidak diubah.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAB"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22

    BIRU     = "1F4E79"
    BIRU_MUD = "BDD7EE"
    ABU      = "F2F2F2"
    KUNING   = "FFF2CC"
    HIJAU    = "E2EFDA"
    PUTIH    = "FFFFFF"

    def _cell(row, col, val, bold=False, italic=False, color="000000",
              bg=None, align="left", fmt=None, border=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=10, bold=bold, italic=italic, color=color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if bg:
            c.fill = PatternFill(fill_type="solid", fgColor=bg)
        if fmt:
            c.number_format = fmt
        if border:
            c.border = BORDER_TIPIS
        return c

    # ── Baris 1–3: Metadata proyek ───────────────────────────────────
    meta_rows = [
        ("Nama Proyek",    "Ganti teks ini dengan nama proyek Anda"),
        ("Lokasi",         "Kabupaten/Kota, Provinsi"),
        ("Tahun Anggaran", 2025),
    ]
    for i, (label, contoh) in enumerate(meta_rows, start=1):
        ws.row_dimensions[i].height = 18
        _cell(i, 1, label, bold=True, bg=ABU)
        # Merge kolom B–F untuk nilai
        ws.merge_cells(f"B{i}:F{i}")
        c = ws.cell(row=i, column=2, value=contoh)
        c.font = Font(name="Calibri", size=10,
                      italic=(i < 3), color="595959" if i < 3 else "000000")
        c.border = BORDER_TIPIS

    # ── Baris 4: kosong (agar parser tidak salah deteksi) ────────────
    ws.row_dimensions[4].height = 8

    # ── Baris 5: Header tabel (JANGAN DIUBAH nama kolomnya) ──────────
    headers = [
        ("No",            "center"),
        ("Uraian Pekerjaan", "left"),
        ("Satuan",        "center"),
        ("Volume",        "center"),
        ("Harga Satuan",  "right"),
        ("Jumlah",        "right"),
    ]
    ws.row_dimensions[5].height = 20
    for col_idx, (h, align) in enumerate(headers, start=1):
        c = _cell(5, col_idx, h, bold=True, color=PUTIH, bg=BIRU, align=align)
        c.font = Font(name="Calibri", size=10, bold=True, color=PUTIH)

    # ── Baris 6+: Contoh data ────────────────────────────────────────
    # Format baris divisi: kolom A=None, B=nama kelompok (huruf besar/romawi),
    #   kolom C–F dikosongkan. Parser mendeteksi ini sebagai group header.
    # Format baris item: isi semua kolom.
    contoh = [
        # (no, uraian, satuan, volume, harga_satuan, jumlah)
        (None, "I. PEKERJAAN PERSIAPAN",            None,    None,       None,        None),
        (1,    "Pembersihan dan Perataan Lahan",     "m2",    500,        5000,        2500000),
        (2,    "Pengukuran dan Pemasangan Bouwplank","m1",    80,         15000,       1200000),
        (3,    "Pembuatan Direksi Keet",             "unit",  1,          5000000,     5000000),
        (None, "II. PEKERJAAN TANAH",               None,    None,       None,        None),
        (4,    "Galian Tanah untuk Pondasi",         "m3",    120,        45000,       5400000),
        (5,    "Urugan Kembali Tanah Bekas Galian",  "m3",    60,         25000,       1500000),
        (6,    "Urugan Pasir Bawah Lantai",          "m3",    15,         180000,      2700000),
        (None, "III. PEKERJAAN PONDASI",            None,    None,       None,        None),
        (7,    "Pondasi Batu Kali 1:4",              "m3",    80,         850000,      68000000),
        (8,    "Sloof Beton 15/20 K-175",            "m3",    12,         2500000,     30000000),
        (None, "IV. PEKERJAAN STRUKTUR",            None,    None,       None,        None),
        (9,    "Kolom Beton 20/20 K-175",            "m3",    18,         3200000,     57600000),
        (10,   "Balok Latei 10/15 K-175",            "m3",    8,          3000000,     24000000),
    ]

    baris = 6
    for no, uraian, satuan, volume, harga, jumlah in contoh:
        ws.row_dimensions[baris].height = 17
        if no is None:
            # Baris kelompok / divisi
            ws.merge_cells(f"A{baris}:F{baris}")
            c = ws.cell(row=baris, column=1, value=uraian)
            c.font = Font(name="Calibri", size=10, bold=True)
            c.fill = PatternFill(fill_type="solid", fgColor=BIRU_MUD)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = BORDER_TIPIS
        else:
            vals = [no, uraian, satuan, volume, harga, jumlah]
            aligns = ["center", "left", "center", "center", "right", "right"]
            fmts   = [None, None, None, "#,##0.##", FORMAT_RUPIAH, FORMAT_RUPIAH]
            for ci, (val, al, fmt) in enumerate(zip(vals, aligns, fmts), start=1):
                _cell(baris, ci, val, align=al, fmt=fmt)
        baris += 1

    # ── Baris kosong ────────────────────────────────────────────────
    baris += 1

    # ── Tambah baris kosong untuk input user (30 baris kosong) ──────
    ws.row_dimensions[baris].height = 8
    baris += 1
    for _ in range(30):
        ws.row_dimensions[baris].height = 17
        for ci in range(1, 7):
            _cell(baris, ci, None,
                  bg=HIJAU if ci == 2 else None,
                  fmt=FORMAT_RUPIAH if ci in [5, 6] else ("#,##0.##" if ci == 4 else None))
        baris += 1

    # Freeze panes di baris header
    ws.freeze_panes = "A6"

    # ── Sheet PETUNJUK ───────────────────────────────────────────────
    ws2 = wb.create_sheet("PETUNJUK")
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 55

    def _p(row, col, val, bold=False, bg=None, color="000000"):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=10, bold=bold, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bg:
            c.fill = PatternFill(fill_type="solid", fgColor=bg)
        return c

    r = 1
    ws2.merge_cells(f"A{r}:C{r}")
    c = ws2.cell(row=r, column=1, value="PETUNJUK PENGISIAN TEMPLATE RAB")
    c.font = Font(name="Calibri", size=13, bold=True, color=PUTIH)
    c.fill = PatternFill(fill_type="solid", fgColor=BIRU)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r].height = 26
    r += 2

    seksi = [
        ("INFO PROYEK (baris 1-3 di sheet RAB)", [
            ("Nama Proyek",    "Nama lengkap proyek. Contoh: Pembangunan Gedung Kantor Desa"),
            ("Lokasi",         "Kabupaten/Kota diikuti Provinsi. Contoh: Kab. Bogor, Jawa Barat"),
            ("Tahun Anggaran", "Tahun 4 digit. Contoh: 2024, 2025"),
        ]),
        ("KOLOM TABEL (baris 5 adalah header — JANGAN diubah namanya)", [
            ("No",           "Nomor urut item (opsional, boleh kosong)"),
            ("Uraian Pekerjaan", "Nama item pekerjaan. WAJIB diisi di setiap baris"),
            ("Satuan",       "Satuan ukur: m2, m3, m1, unit, kg, ls, titik, dll."),
            ("Volume",       "Kuantitas/jumlah pekerjaan. Gunakan angka biasa (bukan Rp)"),
            ("Harga Satuan", "Harga per satuan dalam Rupiah. Boleh dengan titik ribuan"),
            ("Jumlah",       "Opsional — sistem menghitung ulang otomatis (Volume x Harga)"),
        ]),
        ("BARIS KELOMPOK / DIVISI", [
            ("Format",    "Isi kolom Uraian Pekerjaan saja, kosongkan kolom lain"),
            ("Contoh",    "I. PEKERJAAN PERSIAPAN  atau  A. PEKERJAAN TANAH"),
            ("Deteksi",   "Sistem mendeteksi otomatis jika ada angka Romawi atau huruf besar semua"),
        ]),
        ("TIPS PENTING", [
            ("Jangan merge cell",  "Di area tabel data (baris 5 ke bawah), jangan gabungkan sel"),
            ("Format angka",       "Tulis angka biasa: 850000 atau 850.000 atau Rp 850.000"),
            ("Desimal",            "Gunakan koma atau titik: 12.5 atau 12,5 keduanya diterima"),
            ("Baris kosong",       "Boleh ada baris kosong di tengah, sistem akan melewatinya"),
            ("Kolom tambahan",     "Kolom di luar A-F diabaikan, aman untuk catatan tambahan"),
        ]),
    ]

    for judul, items in seksi:
        ws2.merge_cells(f"A{r}:C{r}")
        c2 = ws2.cell(row=r, column=1, value=judul)
        c2.font = Font(name="Calibri", size=10, bold=True, color=PUTIH)
        c2.fill = PatternFill(fill_type="solid", fgColor="2E75B6")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[r].height = 18
        r += 1
        for label, keterangan in items:
            _p(r, 2, label, bold=True, bg=ABU)
            _p(r, 3, keterangan)
            ws2.row_dimensions[r].height = 18
            r += 1
        r += 1

    path_template = TEMPLATE_DIR / "template_rab.xlsx"
    wb.save(str(path_template))
    return str(path_template)
